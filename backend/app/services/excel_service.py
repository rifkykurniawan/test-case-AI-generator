import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.exceptions import ExportError
from app.schemas.response import GenerateResponse


class ExcelService:
    """Service to export generated test cases to an Excel workbook using openpyxl."""

    def __init__(self):
        # Color palette styles (sleek premium blue theme)
        self.header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        self.header_fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid"
        )

        self.title_font = Font(name="Segoe UI", size=14, bold=True, color="1F4E78")
        self.bold_font = Font(name="Segoe UI", size=10, bold=True)
        self.regular_font = Font(name="Segoe UI", size=10)

        self.border_thin = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
        self.align_top_left = Alignment(
            horizontal="left", vertical="top", wrap_text=True
        )
        self.align_center = Alignment(horizontal="center", vertical="top")

    def _auto_fit_columns(self, ws, max_width_limit: int = 50):
        """Helper to fit column widths automatically with padding."""
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or "")
                # Handle multi-line cells length estimation
                lines = val.split("\n")
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
            # Give padding
            ws.column_dimensions[col_letter].width = min(
                max(max_len + 3, 12), max_width_limit
            )

    def export_to_bytes(self, requirement_input: str, data: GenerateResponse) -> bytes:
        """Generates an Excel workbook bytes representing the test generation output.

        Args:
            requirement_input: The raw user-supplied requirement description.
            data: The structured GenerateResponse object.

        Returns:
            bytes: Binary content of the Excel sheet.

        Raises:
            ExportError: If workbook generation fails.
        """
        try:
            wb = Workbook()

            # --- Sheet 1: Test Cases ---
            ws_tc = wb.active
            ws_tc.title = "Test Cases"
            ws_tc.views.sheetView[0].showGridLines = True

            ws_tc.cell(
                row=2, column=2, value="Functional Test Cases"
            ).font = self.title_font

            headers = [
                "ID",
                "Title",
            ]
            for col_idx, header in enumerate(headers, 2):
                cell = ws_tc.cell(row=4, column=col_idx, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.align_center

            row_idx = 5
            for tc in data.testCases:
                cells = [
                    ws_tc.cell(row=row_idx, column=2, value=tc.id),
                    ws_tc.cell(row=row_idx, column=3, value=tc.title),
                ]

                for c in cells:
                    c.font = self.regular_font
                    c.border = self.border_thin
                    c.alignment = self.align_top_left

                # Center ID
                cells[0].alignment = self.align_center
                row_idx += 1

            self._auto_fit_columns(ws_tc, max_width_limit=45)

            # Write to buffer
            file_stream = io.BytesIO()
            wb.save(file_stream)
            file_stream.seek(0)
            return file_stream.getvalue()

        except Exception as e:
            raise ExportError(f"Excel file generation failed: {e}")


stream = io.BytesIO()
